# run_benchmark.py
# One-shot: scrape -> dumps/fees_<bank>.txt -> parse (strict + fallback) -> build template -> fill -> save Benchmark_Results.xlsx

# ───────────────────── BEGIN: YOUR SCRAPER (UNCHANGED) ──────────────────────
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import re

URL = "https://www.bankacilikurunvehizmetucretleri.org.tr/bireysel-ucret/liste"

list_of_banks = [
    "Akbank T.A.Ş.",
    "Türkiye Garanti Bankası A.Ş.",
    "Türkiye İş Bankası A.Ş.",
    "Yapı ve Kredi Bankası A.Ş.",
    "Türkiye Cumhuriyeti Ziraat Bankası A.Ş.",
    "Türkiye Halk Bankası A.Ş.",
    "Türkiye Vakıflar Bankası T.A.O.",
    "Denizbank A.Ş.",
    "QNB Bank A.Ş.",
]

def safe_name(s: str) -> str:
    return re.sub(r"[^\w\-\.]+", "_", s.strip())

with sync_playwright() as pw:
    browser = pw.chromium.launch(headless=True)
    page = browser.new_page()

    page.goto(URL, wait_until="load")
    page.wait_for_load_state("networkidle")

    for BANK_LABEL in list_of_banks:
        outfile = f"fees_{safe_name(BANK_LABEL)}.txt"
        with open(outfile, "w", encoding="utf-8") as out:
            out.write(f"=== BANK: {BANK_LABEL} ===\n")

            try:
                page.select_option("#bankList", label=BANK_LABEL)
            except Exception:
                page.locator('button[data-id="bankList"]').click()
                page.locator('.dropdown-menu.show .dropdown-item .text', has_text=BANK_LABEL).click()
            page.wait_for_load_state("networkidle")

            tab_hrefs = page.eval_on_selector_all(
                'ul.nav-tabs a[role="tab"]',
                'els => els.map(e => e.getAttribute("href")).filter(Boolean)'
            )

            for href in tab_hrefs:
                tab = page.locator(f'a[role="tab"][href="{href}"]')
                tab.scroll_into_view_if_needed()
                tab.click()
                page.wait_for_load_state("networkidle")
                page.locator(href).wait_for(state="visible", timeout=15000)

                page.evaluate(f"""
                document.querySelectorAll('{href} .collapse').forEach(el => {{
                    el.classList.add('show'); el.style.height='auto';
                }});
                """)
                togglers = page.locator(
                    f'{href} [data-toggle="collapse"], '
                    f'{href} [data-bs-toggle="collapse"], '
                    f'{href} .card-header button[aria-controls]'
                )
                for i in range(togglers.count()):
                    t = togglers.nth(i)
                    exp = t.get_attribute("aria-expanded")
                    if exp is None or exp.lower() == "false":
                        t.scroll_into_view_if_needed()
                        t.click()
                        page.wait_for_load_state("networkidle")

                selects = page.locator(f"{href} select")
                for i in range(selects.count()):
                    sel = selects.nth(i)
                    labels = [s.strip() for s in sel.locator("option").all_text_contents()]
                    choice = None
                    for lab in labels:
                        if lab and lab.lower() not in ("hepsi", "seçiniz", "seciniz", "tümü"):
                            choice = lab
                            break
                    if choice:
                        try:
                            sel.select_option(label=choice)
                            page.wait_for_load_state("networkidle")
                        except Exception:
                            pass

                try:
                    page.locator(f"{href} tbody tr").first.wait_for(state="visible", timeout=3000)
                except Exception:
                    pass

                pane_html = page.locator(href).inner_html()
                soup = BeautifulSoup(pane_html, "html.parser")

                table_items = soup.select(".table_item")
                if not table_items:
                    table_items = [soup]

                try:
                    tab_title = tab.inner_text().strip()
                except Exception:
                    tab_title = href

                out.write(f"\n\n===== TAB: {tab_title} ({href}) =====\n")

                any_table = False
                for block in table_items:
                    section_title = None
                    head = block.find(["h3", "h4", "h5", "h6"])
                    if head:
                        section_title = head.get_text(" ", strip=True)

                    cards = block.select(".card")
                    for card in cards:
                        sub_heading = ""
                        ch = card.select_one(".card-header")
                        if ch:
                            sub_heading = ch.get_text(" ", strip=True)

                        tables = card.select(".card-body table")
                        for ti, tbl in enumerate(tables, 1):
                            any_table = True
                            out.write(f"\n--- TABLE {ti} ---\n")
                            out.write(f"SECTION: {section_title or '(no section title)'}\n")
                            out.write(f"SUB-HEADING: {sub_heading or '(no sub-heading)'}\n")

                            headers = [th.get_text(" ", strip=True) for th in tbl.select("thead th")]
                            if headers:
                                out.write("HEADERS: " + " | ".join(headers) + "\n")

                            for row in tbl.select("tbody tr"):
                                cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                                if any(cells):
                                    out.write(" | ".join(cells) + "\n")

                if not any_table:
                    tables = soup.select("table")
                    out.write(f"\n(no .card tables found; fallback tables: {len(tables)})\n")
                    for ti, tbl in enumerate(tables, 1):
                        out.write(f"\n--- TABLE {ti} ---\n")
                        for row in tbl.select("tbody tr"):
                            cells = [c.get_text(" ", strip=True) for c in row.find_all(["th", "td"])]
                            if any(cells):
                                out.write(" | ".join(cells) + "\n")

                page.wait_for_timeout(250)

    browser.close()

print("Done. Created one .txt file per bank in the current folder.")
# ────────────────────── END: YOUR SCRAPER (UNCHANGED) ───────────────────────


# ─────────────── BEGIN: YOUR EXCEL TEMPLATE FUNCTION (UNCHANGED) ─────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def build_benchmark_template():
    year = datetime.now().year
    benchmark_title = f"BENCHMARKING-{year}"

    bank_headers = ["GARANTI", "AKBANK", "İŞBANKASI", "YKB", "ZİRAAT",
                    "HALKBANK", "VAKIFBANK", "DENIZBANK", "FINASNBANK"]

    bank_colors = {
        "GARANTI": "007A33",
        "AKBANK": "E60012",
        "İŞBANKASI": "003A8C",
        "YKB": "1D2E5A",
        "ZİRAAT": "E30613",
        "HALKBANK": "005DAA",
        "VAKIFBANK": "FFB81C",
        "DENIZBANK": "004C97",
        "FINASNBANK": "5C1E4F",
    }

    sections = [
        ("ŞANS OYUNLARI", [""]),
        ("EFT", [
            "HESAPTAN EFT - Şube",
            "HESAPTAN EFT - ATM",
            "HESAPTAN EFT - Mobil",
            "DÜZENLİ EFT",
            "KREDİ KARTINDAN FATURA ÖDEME",
        ]),
        ("HAVALE", [
            "HESAPTAN HAVALE - Şube",
            "HESAPTAN HAVALE - ATM",
            "HESAPTAN HAVALE - Mobil",
            "DÜZENLİ HAVALE",
        ]),
        ("SWIFT", [
            "GİDEN SWIFT",
            "GELEN SWIFT",
            "GİDEN SWIFT - Mobil",
        ]),
        ("ÇEK", [
            "ÇEK TAHSİLİ BAŞKA BANKA",
            "ÇEK TAHSİLİ GB",
            "AYNI ŞUBE ÇEK TAHSİLATI",
            "BAŞKA ŞUBE ÇEK TAHSİLATI",
            "BLOKE ÇEK ÖDEME",
            "ÇEK İADE",
            "BLOKE ÇEK DÜZENLEME",
            "YP ÇEK TAKASA GÖNDERME",
            "ÇEK KARNESİ SAYFA ÜCRETİ",
        ]),
        ("SENET", [
            "SENET TAHSİLE ALMA",
            "MUAMELESİZ SENET İADESİ",
        ]),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "BENCHMARK"

    widths = {1: 24, 2: 46}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    for col_idx in range(3, 3 + len(bank_headers)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    black_bold = Font(bold=True, color="000000")

    headers = [benchmark_title, ""] + bank_headers
    ws.append(headers)

    c = ws.cell(row=1, column=1); c.font = bold; c.alignment = center; c.border = border_all; c.fill = PatternFill("solid", fgColor="EEEEEE")
    c = ws.cell(row=1, column=2); c.border = border_all; c.fill = PatternFill("solid", fgColor="DDDDDD")

    for i, bank in enumerate(bank_headers, start=3):
        c = ws.cell(row=1, column=i, value=bank)
        fill_color = bank_colors.get(bank, "333333")
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.font = black_bold if bank == "VAKIFBANK" else white_bold
        c.alignment = center
        c.border = border_all

    row_cursor = 2
    for main_section, sub_items in sections:
        start_row = row_cursor
        end_row = row_cursor + len(sub_items) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        main_cell = ws.cell(row=start_row, column=1, value=main_section)
        main_cell.font = bold; main_cell.alignment = center; main_cell.border = border_all

        for sub in sub_items:
            ws.cell(row=row_cursor, column=2, value=sub).alignment = left
            ws.cell(row=row_cursor, column=2).border = border_all
            for col in range(3, 3 + len(bank_headers)):
                cell_to_style = ws.cell(row=row_cursor, column=col)
                cell_to_style.border = border_all
                cell_to_style.alignment = center
            row_cursor += 1

    return wb
# ──────────────── END: YOUR EXCEL TEMPLATE FUNCTION (UNCHANGED) ──────────────


# ─────────── BEGIN: ORIGINAL STRICT PARSER + FALLBACK (TL/TRY tolerant) ──────
import re as _re
from pathlib import Path as _Path

TL = "TL"

def norm_money(s: str) -> str:
    if not s:
        return ""
    s = s.replace("TRY", TL).replace("\xa0", " ")
    s = _re.sub(r"\s+", " ", s.strip())
    return s

def first_group(pat, text, flags=_re.S):
    m = _re.search(pat, text, flags)
    return m.group(1).strip() if m else ""

def amount_from_line(text, label_key):
    m = _re.search(rf"{_re.escape(label_key)}.*?(\d[\d\.\,]*\s*TRY)", text, _re.S)
    return norm_money(m.group(1)) if m else ""

def all_amounts_on_line(text, label_key):
    line = first_group(rf"({_re.escape(label_key)}.*?$)", text, flags=_re.M)
    if not line:
        return ""
    vals = [norm_money(v) for v in _re.findall(r"(\d[\d\.,]*\s*TRY|\d[\d\.,]*\s*USD)", line)]
    return " / ".join(vals)

def percent_from_line(text, label_key):
    m = _re.search(rf"{_re.escape(label_key)}.*?% ?(\d+(?:,\d+)?)", text, _re.S | _re.I)
    return f"%{m.group(1)}" if m else ""

def three_band_from_block(text, heading_key, channel):
    block_pat = rf"{_re.escape(heading_key)}.*?\|\s*{channel}\s*\|.*?(?=(\n---|\Z))"
    block = _re.search(block_pat, text, _re.S)
    if not block:
        return ""
    seg = block.group(0)
    low  = first_group(r"1\s*TRY\s*-\s*6\.?300\s*TRY.*?\|\s*([\d\.,]+\s*TRY)", seg)
    mid  = first_group(r"6\.?300,?01\s*TRY\s*-\s*304\.?800\s*TRY.*?\|\s*([\d\.,]+\s*TRY)", seg)
    high = first_group(r"304\.?800,?01\s*TRY.*?\|\s*([\d\.,]+\s*TRY)", seg)
    parts = [norm_money(x) for x in (low, mid, high) if x]
    return " - ".join(parts)

def parse_dump(path: str) -> dict:
    t = _Path(path).read_text(encoding="utf-8", errors="ignore")
    out = {}

    # ŞANS OYUNLARI
    s_min = first_group(r"Şans Oyunu Ödemeleri Aracılık.*?Asgari Tutar.*?(\d[\d\.,]*\s*TRY)", t)
    s_max = first_group(r"Şans Oyunu Ödemeleri Aracılık.*?Azami Tutar.*?(\d[\d\.,]*\s*TRY)", t)
    so = f"{norm_money(s_min)} - {norm_money(s_max)}".strip(" -")
    if so: out["ŞANS OYUNLARI"] = so

    # EFT
    out["HESAPTAN EFT - Şube"]  = three_band_from_block(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - EFT Gönderimi", "Şube")
    out["HESAPTAN EFT - ATM"]   = three_band_from_block(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - EFT Gönderimi", "ATM")
    out["HESAPTAN EFT - Mobil"] = three_band_from_block(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - EFT Gönderimi", "Mobil Kanal")
    out["DÜZENLİ EFT"] = three_band_from_block(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli EFT Gönderimi", "İnternet") \
                      or three_band_from_block(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli EFT Gönderimi", "Mobil Kanal")
    out["KREDİ KARTINDAN FATURA ÖDEME"] = percent_from_line(t, "Fatura Ödeme / Kurum Ödeme - Düzenli Ödemeler")

    # HAVALE
    base = "Havale Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Havale Gönderimi"
    out["HESAPTAN HAVALE - Şube"]  = three_band_from_block(t, base, "Şube")
    out["HESAPTAN HAVALE - ATM"]   = three_band_from_block(t, base, "ATM")
    out["HESAPTAN HAVALE - Mobil"] = three_band_from_block(t, base, "Mobil Kanal")
    out["DÜZENLİ HAVALE"] = three_band_from_block(t, "Havale Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli Havale Gönderimi", "İnternet") \
                         or three_band_from_block(t, "Havale Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli Havale Gönderimi", "Mobil Kanal")

    # SWIFT
    wu_min = first_group(r"Western Union.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*USD)", t)
    wu_max = first_group(r"Western Union.*?Asgari Tutar\s*\|\s*Azami Tutar.*?USD\s*\|\s*([\d\.,]+\s*USD)", t)
    dig_min = first_group(r"Diğer Aracı Firmalar.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*TRY)", t)
    dig_max = first_group(r"Diğer Aracı Firmalar.*?Asgari Tutar\s*\|\s*Azami Tutar.*?TRY\s*\|\s*([\d\.,]+\s*TRY)", t)
    parts = []
    if wu_min or wu_max: parts.append(f"WU: {norm_money(wu_min)}–{norm_money(wu_max)}")
    if dig_min or dig_max: parts.append(f"Diğer: {norm_money(dig_min)}–{norm_money(dig_max)}")
    out["GİDEN SWIFT"] = "; ".join([p for p in parts if p])

    g_min = first_group(r"Uluslararası Para Transferi Ödenmesi.*?Hesaba Gelen - Yurtdışı Bankadan.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*TRY)", t)
    g_max = first_group(r"Uluslararası Para Transferi Ödenmesi.*?Hesaba Gelen - Yurtdışı Bankadan.*?Asgari Tutar\s*\|\s*Azami Tutar.*?TRY\s*\|\s*([\d\.,]+\s*TRY)", t)
    if g_min or g_max:
        out["GELEN SWIFT"] = f"Hesaba: Asgari {norm_money(g_min)} | Azami {norm_money(g_max)}".strip(" |")

    mob_swift_block = _re.search(r"Uluslararası Para transferi.*?Hesaptan - Hesaba.*?Mobil Kanal.*?(?=\n---|\Z)", t, _re.S)
    if mob_swift_block:
        seg = mob_swift_block.group(0)
        amounts = [norm_money(m) for m in _re.findall(r"(\d[\d\.,]+\s*TRY)", seg)]
        if len(amounts) >= 2:
            out["GİDEN SWIFT - Mobil"] = f"{amounts[0]} - {amounts[1]}"

    # ÇEK
    cek_tahsilat_block = first_group(r"(SUB-HEADING: Çek Tahsilat Ücreti.*?)(?=\n---|\Z)", t)
    out["ÇEK TAHSİLİ BAŞKA BANKA"] = f"{percent_from_line(cek_tahsilat_block, 'Diğer Banka Çeki -')} Asgari Tutar: {amount_from_line(cek_tahsilat_block, 'Diğer Banka Çeki -')} Azami Tutar: {all_amounts_on_line(cek_tahsilat_block, 'Diğer Banka Çeki -')}"
    out["ÇEK TAHSİLİ GB"] = percent_from_line(t, "Çek Tahsili GB") or percent_from_line(t, "Çek Tahsili G.B")
    out["AYNI ŞUBE ÇEK TAHSİLATI"] = f"{percent_from_line(cek_tahsilat_block, 'Aynı Banka Çeki -')} Asgari Tutar: {amount_from_line(cek_tahsilat_block, 'Aynı Banka Çeki -')} Azami Tutar: {all_amounts_on_line(cek_tahsilat_block, 'Aynı Banka Çeki -')}"
    out["BAŞKA ŞUBE ÇEK TAHSİLATI"] = percent_from_line(t, "Başka Şube Çek Tahsili")
    out["BLOKE ÇEK ÖDEME"] = amount_from_line(t, "Bloke Çek Ödeme")
    out["ÇEK İADE"] = amount_from_line(t, "Çek İade Ücreti")
    out["BLOKE ÇEK DÜZENLEME"] = f"{percent_from_line(t, 'Çek Düzenleme -')} Asgari Tutar: {amount_from_line(t, 'Çek Düzenleme -')} Azami Tutar: {all_amounts_on_line(t, 'Çek Düzenleme -')}"
    out["YP ÇEK TAKASA GÖNDERME"] = f"{percent_from_line(t, 'Döviz Çekleri Tahsilatı (Diğer Banka) -')} Asgari Tutar: {amount_from_line(t, 'Döviz Çekleri Tahsilatı (Diğer Banka) -')} Azami Tutar: {all_amounts_on_line(t, 'Döviz Çekleri Tahsilatı (Diğer Banka) -')}"
    out["ÇEK KARNESİ SAYFA ÜCRETİ"] = amount_from_line(t, "Çek Defteri (Yaprak Başı)")

    # SENET
    out["SENET TAHSİLE ALMA"] = amount_from_line(t, "Aynı Banka Senet Tahsili -") or amount_from_line(t, "Senet Tahsile Alma")
    out["MUAMELESİZ SENET İADESİ"] = amount_from_line(t, "Senet İade Ücreti")

    return out

# Fallback tolerant to TL/TRY and small label variants; only used when strict value is empty
_CCY = r"(?:TL|TRY)"

def _amount_from_line_loose(text, label_key):
    m = _re.search(rf"{_re.escape(label_key)}.*?(\d[\d\.\,]*\s*{_CCY})", text, _re.S | _re.I)
    return norm_money(m.group(1)) if m else ""

def _three_band_loose(text, heading_keys, channel_aliases):
    for hk in heading_keys:
        for ch in channel_aliases:
            block = _re.search(rf"{_re.escape(hk)}.*?\|\s*{ch}\s*\|.*?(?=(\n---|\Z))", text, _re.S | _re.I)
            if not block:
                continue
            seg = block.group(0)
            low  = first_group(rf"1\s*{_CCY}\s*-\s*6\.?300\s*{_CCY}.*?\|\s*([\d\.,]+\s*{_CCY})", seg)
            mid  = first_group(rf"6\.?300,?01\s*{_CCY}\s*-\s*304\.?800\s*{_CCY}.*?\|\s*([\d\.,]+\s*{_CCY})", seg)
            high = first_group(rf"304\.?800,?01\s*{_CCY}.*?\|\s*([\d\.,]+\s*{_CCY})", seg)
            parts = [norm_money(x) for x in (low, mid, high) if x]
            if parts:
                return " - ".join(parts)
    return ""

# ── NEW ultra-tolerant helpers (added; used only when strict+loose are empty)
def _three_band_generic(text, heading_keys, channel_aliases):
    """
    Very tolerant: after locating a heading + channel, grab the first three
    price cells from the 3rd column of subsequent rows (no hard-coded ranges).
    """
    for hk in heading_keys:
        m = _re.search(rf"{_re.escape(hk)}", text, _re.I)
        if not m:
            continue
        tail = text[m.end():]

        chan_match = None
        for ch in channel_aliases:
            chan_match = _re.search(rf"^\s*[^\n]*\|\s*{ch}\s*\|", tail, _re.M | _re.I)
            if chan_match:
                break
        if not chan_match:
            continue

        seg = tail[chan_match.end():]
        stop = _re.search(
            rf"^\s*[^\n]*\|\s*(Şube|ATM|İnternet|Internet|Mobil Kanal|Mobil|Çağrı Merkezi|Dijital Kanallar|Dijital)\s*\|"
            rf"|{_re.escape(hk)}", seg, _re.M | _re.I
        )
        if stop:
            seg = seg[:stop.start()]

        vals = _re.findall(rf"^\s*[^\n]*\|\s*[^\n]*\|\s*([\d\.,]+\s*{_CCY})", seg, _re.M | _re.I)
        vals = [norm_money(v) for v in vals[:3]]
        if vals:
            return " - ".join(vals)
    return ""

def _sans_oyunlari_loose(text):
    """
    Tolerant Şans Oyunları finder:
    - prefers SUB-HEADING blocks
    - falls back to any paragraph containing Şans Oyun and Asgari/Azami
    """
    block = first_group(r"(SUB-HEADING:\s*Şans[^\n]*?)(?=\n---|\Z)", text, flags=_re.S | _re.I)
    if not block:
        block = first_group(r"(Şans\s*Oyun[^\n]*?Asgari.*?Azami.*?)(?=\n---|\Z)", text, flags=_re.S | _re.I)
    if not block:
        return ""
    s_min = first_group(rf"Asgari Tutar.*?\|\s*Azami Tutar.*?\|\s*([\d\.,]+\s*{_CCY})", block, flags=_re.S | _re.I)
    s_max = first_group(rf"Azami Tutar.*?\|\s*([\d\.,]+\s*{_CCY})", block, flags=_re.S | _re.I)
    if s_min or s_max:
        return f"{norm_money(s_min)} - {norm_money(s_max)}".strip(" -")
    return ""

def parse_dump_with_fallback(path: str) -> dict:
    base = parse_dump(path).copy()
    t = _Path(path).read_text(encoding="utf-8", errors="ignore")
    t = (t.replace("\u00a0", " ")
           .replace("–", "-").replace("—", "-"))

    def need(k): return not base.get(k)

    # NEW: rescue ŞANS OYUNLARI (e.g., İşbankası variants)
    if need("ŞANS OYUNLARI"):
        base["ŞANS OYUNLARI"] = _sans_oyunlari_loose(t) or base.get("ŞANS OYUNLARI", "")

    eft_heads = [
        "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - EFT Gönderimi",
        "EFT Gönderilmesi",
    ]
    ch_mobil = ["Mobil Kanal", "Mobil"]

    if need("HESAPTAN EFT - Şube"):
        base["HESAPTAN EFT - Şube"] = _three_band_loose(t, eft_heads, ["Şube"])
    if need("HESAPTAN EFT - ATM"):
        base["HESAPTAN EFT - ATM"] = _three_band_loose(t, eft_heads, ["ATM"])
    if need("HESAPTAN EFT - Mobil"):
        base["HESAPTAN EFT - Mobil"] = _three_band_loose(t, eft_heads, ch_mobil)
    if need("DÜZENLİ EFT"):
        base["DÜZENLİ EFT"] = (_three_band_loose(
            t,
            ["EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli EFT Gönderimi",
             "Düzenli EFT Gönderimi"],
            ["İnternet", "Internet"] + ch_mobil
        ) or base.get("DÜZENLİ EFT", ""))

    # NEW: ultra-tolerant Düzenli EFT rescue (e.g., YKB "Dijital Kanallar")
    if need("DÜZENLİ EFT"):
        base["DÜZENLİ EFT"] = (_three_band_generic(
            t,
            [
                "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli EFT Gönderimi",
                "Düzenli EFT Gönderimi",
                "Düzenli EFT"
            ],
            ["İnternet", "Internet", "Mobil Kanal", "Mobil",
             "İnternet/Mobil", "İnternet - Mobil",
             "Dijital Kanallar", "Dijital"]
        ) or base.get("DÜZENLİ EFT", ""))

    hav_heads = [
        "Havale Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Havale Gönderimi",
        "Havale Gönderilmesi",
    ]
    if need("HESAPTAN HAVALE - Şube"):
        base["HESAPTAN HAVALE - Şube"] = _three_band_loose(t, hav_heads, ["Şube"])
    if need("HESAPTAN HAVALE - ATM"):
        base["HESAPTAN HAVALE - ATM"] = _three_band_loose(t, hav_heads, ["ATM"])
    if need("HESAPTAN HAVALE - Mobil"):
        base["HESAPTAN HAVALE - Mobil"] = _three_band_loose(t, hav_heads, ch_mobil)
    if need("DÜZENLİ HAVALE"):
        base["DÜZENLİ HAVALE"] = (_three_band_loose(
            t,
            ["Havale Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli Havale Gönderimi",
             "Düzenli Havale Gönderimi"],
            ["İnternet", "Internet"] + ch_mobil
        ) or base.get("DÜZENLİ HAVALE", ""))

    singles = [
        ("BLOKE ÇEK ÖDEME", "Bloke Çek Ödeme"),
        ("ÇEK İADE", "Çek İade Ücreti"),
        ("ÇEK KARNESİ SAYFA ÜCRETİ", "Çek Defteri (Yaprak Başı)"),
        ("SENET TAHSİLE ALMA", "Senet Tahsile Alma"),
        ("MUAMELESİZ SENET İADESİ", "Senet İade Ücreti"),
    ]
    for key, label in singles:
        if need(key):
            base[key] = _amount_from_line_loose(t, label)

    if need("GİDEN SWIFT") and _re.search(r"Western Union|Diğer Aracı", t, _re.I):
        wu_min = first_group(rf"Western Union.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*(?:USD|{_CCY}))", t)
        wu_max = first_group(rf"Western Union.*?Asgari Tutar\s*\|\s*Azami Tutar.*?(?:USD|{_CCY})\s*\|\s*([\d\.,]+\s*(?:USD|{_CCY}))", t)
        di_min = first_group(rf"Diğer Aracı Firmalar.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*{_CCY})", t)
        di_max = first_group(rf"Diğer Aracı Firmalar.*?Asgari Tutar\s*\|\s*Azami Tutar.*?{_CCY}\s*\|\s*([\d\.,]+\s*{_CCY})", t)
        parts = []
        if wu_min or wu_max: parts.append(f"WU: {norm_money(wu_min)}–{norm_money(wu_max)}")
        if di_min or di_max: parts.append(f"Diğer: {norm_money(di_min)}–{norm_money(di_max)}")
        base["GİDEN SWIFT"] = "; ".join([p for p in parts if p])

    if need("GELEN SWIFT"):
        g_min = first_group(rf"Uluslararası Para Transferi Ödenmesi.*?Hesaba Gelen - Yurtdışı Bankadan.*?Asgari Tutar\s*\|\s*Azami Tutar.*?([\d\.,]+\s*{_CCY})", t)
        g_max = first_group(rf"Uluslararası Para Transferi Ödenmesi.*?Hesaba Gelen - Yurtdışı Bankadan.*?Asgari Tutar\s*\|\s*Azami Tutar.*?{_CCY}\s*\|\s*([\d\.,]+\s*{_CCY})", t)
        if g_min or g_max:
            base["GELEN SWIFT"] = f"Hesaba: Asgari {norm_money(g_min)} | Azami {norm_money(g_max)}".strip(" |")

    return base
# ──────────── END: ORIGINAL STRICT PARSER + FALLBACK ─────────────────────────


# ──────────────── ZİRAAT-SPECIFIC FILTER (ADDED, NO CHANGES TO ABOVE) ────────
import re as _zr
from pathlib import Path as _ZPath

# keep your Ziraat helper behavior (TRY formatting)
_TLZ = "TRY"

def _norm_money_Z(s: str) -> str:
    if not s:
        return ""
    s = s.replace("TL", _TLZ).replace("\xa0", " ")
    s = _zr.sub(r"\s+", " ", s.strip())
    return s

def _first_group_Z(pat, text, flags=_zr.S):
    m = _zr.search(pat, text, flags)
    return m.group(1).strip() if m else ""

def _amount_from_line_Z(text, label_key):
    line_pat = rf"^{_zr.escape(label_key)}.*?(\d[\d\.\,]*\s*{_TLZ})"
    m = _zr.search(line_pat, text, _zr.M | _zr.S)
    return _norm_money_Z(m.group(1)) if m else ""

def _all_amounts_on_line_Z(text, label_key):
    line = _first_group_Z(rf"({_zr.escape(label_key)}.*?$)", text, flags=_zr.M)
    if not line:
        return ""
    vals = [_norm_money_Z(v) for v in _zr.findall(rf"(\d[\d\.\,]*\s*{_TLZ}|\d[\d\.\,]*\s*USD)", line)]
    return " / ".join(vals)

def _percent_from_line_Z(text, label_key):
    line_pat = rf"^{_zr.escape(label_key)}.*?%[\s]*([\d\.,]+)"
    m = _zr.search(line_pat, text, _zr.M | _zr.S | _zr.I)
    return f"%{m.group(1)}" if m else ""

def _combined_fee_from_line_Z(text, label_key):
    line_pat = rf"^{_zr.escape(label_key)}\s*\|.*?\|\s*([^\|]+?)\s*\|"
    line_content = _first_group_Z(line_pat, text, _zr.M | _zr.S)
    return _norm_money_Z(line_content) if line_content else ""

def _three_band_from_block_Z(text, heading_key, channel):
    start_pat = rf"{_zr.escape(heading_key)}\s*\|\s*{channel}\s*\|"
    m = _zr.search(start_pat, text)
    if not m:
        return ""
    tail = text[m.end():]
    stop_pat = rf"(?:^\s*1\s*{_TLZ}\s*-\s*|^\s*[\d\.,]+,01\s*{_TLZ}\s*-|\n){_zr.escape(heading_key)}|^\s*[^\n]*\|\s*(Şube|ATM|İnternet|Mobil Kanal|Çağrı Merkezi)\s*\|"
    stop = _zr.search(stop_pat, tail, flags=_zr.M)
    seg = tail[:stop.start()] if stop else tail

    band_pats = [
        rf"^\s*1\s*{_TLZ}\s*-\s*[\d\.,]+\s*{_TLZ}\s*\|\s*\|\s*([\d\.,]+\s*{_TLZ})",
        rf"^\s*[\d\.,]+,01\s*{_TLZ}\s*-\s*[\d\.,]+\s*{_TLZ}\s*\|\s*\|\s*([\d\.,]+\s*{_TLZ})",
        rf"^\s*[\d\.,]+,01\s*{_TLZ}\s*-\s*\|\s*\|\s*([\d\.,]+\s*{_TLZ})",
    ]
    amounts = []
    for pat in band_pats:
        v = _first_group_Z(pat, seg, flags=_zr.M)
        if v:
            amounts.append(_norm_money_Z(v))
    return " - ".join(amounts)

def parse_dump_ziraat(path: str) -> dict:
    t = _ZPath(path).read_text(encoding="utf-8", errors="ignore")
    out = {}

    # ŞANS OYUNLARI
    so_block = _first_group_Z(r"(SUB-HEADING: Şans Oyunu Ödemeleri Aracılık.*?)(?=\n---|\Z)", t)
    s_min = _first_group_Z(r"Asgari Tutar\s*\|\s*Azami Tutar.*?\|\s*([\d\.,]+\s*TRY)", so_block)
    s_max = _first_group_Z(r"Azami Tutar.*?\|\s*([\d\.,]+\s*TRY)", so_block)
    so = f"{_norm_money_Z(s_min)} - {_norm_money_Z(s_max)}".strip(" -")
    if so:
        out["ŞANS OYUNLARI"] = so

    # EFT
    eft_base_heading = "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - EFT Gönderimi"
    out["HESAPTAN EFT - Şube"]  = _three_band_from_block_Z(t, eft_base_heading, "Şube")
    out["HESAPTAN EFT - ATM"]   = _three_band_from_block_Z(t, eft_base_heading, "ATM")
    out["HESAPTAN EFT - Mobil"] = _three_band_from_block_Z(t, eft_base_heading, "Mobil Kanal")
    out["DÜZENLİ EFT"]          = _three_band_from_block_Z(t, "EFT Gönderilmesi - Hesaptan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Düzenli EFT Gönderimi", "Mobil Kanal")

    fatura_block = _first_group_Z(r"(SUB-HEADING: Fatura Ödeme / Kurum Ödeme - Anlık Ödemeler.*?)(?=\n---|\Z)", t)
    if fatura_block:
        fee = _first_group_Z(r"HEADERS:.*?-\s*\|\s*\|\s*([\d\.,]+\s*TRY)", fatura_block)
        out["KREDİ KARTINDAN FATURA ÖDEME"] = f"{fee} (Kredi kartı ile ödemelerde ek olarak nakit avans faizi uygulanır.)"

    # HAVALE
    havale_base_heading = "Havale Gönderilmesi - Hesaptan  / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta - Havale Gönderimi"
    out["HESAPTAN HAVALE - Şube"]  = _three_band_from_block_Z(t, havale_base_heading, "Şube")
    out["HESAPTAN HAVALE - ATM"]   = _three_band_from_block_Z(t, havale_base_heading, "ATM")
    out["HESAPTAN HAVALE - Mobil"] = _three_band_from_block_Z(t, havale_base_heading, "Mobil Kanal")
    out["DÜZENLİ HAVALE"]          = _three_band_from_block_Z(t, "Havale Gönderilmesi - Kasadan / Hesaba-İsme-Kredi Kartına-Banka Kartına-Ön Ödemeli Karta-Cebe - Düzenli Havale Gönderimi", "Şube")

    # SWIFT
    swift_block = _first_group_Z(r"(SUB-HEADING: Uluslararası Para transferi.*?)(?=\n---|\Z)", t)
    if swift_block:
        from_cash = _percent_from_line_Z(swift_block, "Kasadan - Hesaba")
        from_acct = _percent_from_line_Z(swift_block, "Hesaptan - Hesaba")
        from_web = _all_amounts_on_line_Z(swift_block, "Hesaptan - Hesaba | İnternet")
        parts = []
        if from_cash: parts.append(f"Şube (Kasadan): {from_cash}")
        if from_acct: parts.append(f"Şube (Hesaptan): {from_acct}")
        if from_web: parts.append(f"İnternet: {from_web}")
        out["GİDEN SWIFT"] = "; ".join(parts)
        out["GİDEN SWIFT - Mobil"] = from_web

    out["GELEN SWIFT"] = ""

    # ÇEK
    cek_tahsilat_block = _first_group_Z(r"(SUB-HEADING: Çek Tahsilat Ücreti.*?)(?=\n---|\Z)", t)
    out["ÇEK TAHSİLİ BAŞKA BANKA"] = _amount_from_line_Z(cek_tahsilat_block, 'Diğer Banka Çeki -')
    out["ÇEK TAHSİLİ GB"]           = ""
    out["AYNI ŞUBE ÇEK TAHSİLATI"] = _amount_from_line_Z(cek_tahsilat_block, 'Aynı Banka Çeki -')
    out["BAŞKA ŞUBE ÇEK TAHSİLATI"] = ""
    out["BLOKE ÇEK ÖDEME"]         = ""
    out["ÇEK İADE"]                = ""

    cek_duzenleme_block = _first_group_Z(r"(SUB-HEADING: Çek Defteri ve Çek Düzenleme Ücreti.*?)(?=\n---|\Z)", t)
    out["BLOKE ÇEK DÜZENLEME"]     = _percent_from_line_Z(cek_duzenleme_block, 'Çek Düzenleme -')
    out["YP ÇEK TAKASA GÖNDERME"]   = ""
    out["ÇEK KARNESİ SAYFA ÜCRETİ"] = _amount_from_line_Z(cek_duzenleme_block, "Çek Defteri (Yaprak Başı)")

    # SENET
    senet_tahsil_block = _first_group_Z(r"(SUB-HEADING: Senet Tahsile Alma Ücreti.*?)(?=\n---|\Z)", t)
    out["SENET TAHSİLE ALMA"]       = _combined_fee_from_line_Z(senet_tahsil_block, "Aynı Banka Senet Tahsili -")
    out["MUAMELESİZ SENET İADESİ"]  = ""

    return out
# ───────────────────────── END ZİRAAT-SPECIFIC FILTER ─────────────────────────


# ───────────────────────────── NEW: GLUE LOGIC ───────────────────────────────
def _header_col_map(ws):
    m = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if isinstance(v, str):
            m[v.strip()] = col
    return m

def _row_map(ws):
    rm = {}
    for r in range(2, ws.max_row + 1):
        sec = ws.cell(r, 1).value
        sub = ws.cell(r, 2).value or ""
        if sub:
            rm[sub] = r
        elif sec == "ŞANS OYUNLARI":
            rm["ŞANS OYUNLARI"] = r
    return rm

ROW_ORDER = [
    "ŞANS OYUNLARI",
    "HESAPTAN EFT - Şube","HESAPTAN EFT - ATM","HESAPTAN EFT - Mobil","DÜZENLİ EFT",
    "KREDİ KARTINDAN FATURA ÖDEME",
    "HESAPTAN HAVALE - Şube","HESAPTAN HAVALE - ATM","HESAPTAN HAVALE - Mobil","DÜZENLİ HAVALE",
    "GİDEN SWIFT","GELEN SWIFT","GİDEN SWIFT - Mobil",
    "ÇEK TAHSİLİ BAŞKA BANKA","ÇEK TAHSİLİ GB","AYNI ŞUBE ÇEK TAHSİLATI","BAŞKA ŞUBE ÇEK TAHSİLATI",
    "BLOKE ÇEK ÖDEME","ÇEK İADE","BLOKE ÇEK DÜZENLEME","YP ÇEK TAKASA GÖNDERME","ÇEK KARNESİ SAYFA ÜCRETİ",
    "SENET TAHSİLE ALMA","MUAMELESİZ SENET İADESİ"
]

# map your exact scraped labels -> template headers (keep your bank names)
TEMPLATE_BANK_MAP = {
    "Türkiye Garanti Bankası A.Ş.": "GARANTI",
    "Akbank T.A.Ş.": "AKBANK",
    "Türkiye İş Bankası A.Ş.": "İŞBANKASI",
    "Yapı ve Kredi Bankası A.Ş.": "YKB",
    "Türkiye Cumhuriyeti Ziraat Bankası A.Ş.": "ZİRAAT",
    "Türkiye Halk Bankası A.Ş.": "HALKBANK",
    "Türkiye Vakıflar Bankası T.A.O.": "VAKIFBANK",
    "Denizbank A.Ş.": "DENIZBANK",
    "QNB Bank A.Ş.": "FINASNBANK",   # template header is spelled FINASNBANK
}

def _fill_excel_from_dumps(output_path="Benchmark_Results.xlsx"):
    wb = build_benchmark_template()
    ws = wb.active

    hmap = _header_col_map(ws)
    rmap = _row_map(ws)

    for bank_label in list_of_banks:
        header_name = TEMPLATE_BANK_MAP.get(bank_label)
        if not header_name:
            print(f"[WARN] No template header mapping for bank: {bank_label}")
            continue
        col = hmap.get(header_name)
        if not col:
            print(f"[WARN] Header '{header_name}' not found in sheet.")
            continue

        dump_file = f"fees_{safe_name(bank_label)}.txt"
        try:
            values = parse_dump_with_fallback(dump_file)
        except FileNotFoundError:
            print(f"[WARN] Dump not found: {dump_file} (skipping)")
            continue

        # ── overlay Ziraat-specific values ONLY for Ziraat (fill blanks only)
        if bank_label == "Türkiye Cumhuriyeti Ziraat Bankası A.Ş.":
            z_vals = parse_dump_ziraat(dump_file)
            for k, v in z_vals.items():
                if v and not values.get(k):
                    values[k] = v

        # NEW: Print the filtered results BEFORE writing to cells
        print(f"\n[PREVIEW] {bank_label} -> column '{header_name}' (from {dump_file})")
        for key in ROW_ORDER:
            v = values.get(key, "")
            print(f"  {key}: {v if v else '(empty)'}")
        print("[/PREVIEW]\n")

        for key in ROW_ORDER:
            row = rmap.get(key)
            if not row:
                continue
            ws.cell(row=row, column=col).value = values.get(key, "")

        print(f"[OK] Filled column '{header_name}' from {dump_file}")

    wb.save(output_path)
    print(f"[DONE] Saved filled workbook -> {output_path}")

if __name__ == "__main__":
    _fill_excel_from_dumps("Benchmark_Results.xlsx")
